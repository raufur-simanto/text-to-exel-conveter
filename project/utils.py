import sys
from datetime import datetime, timedelta
from collections import defaultdict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


data_length = []

# Function to convert 24-hour time to 12-hour AM/PM format
def convert_time_format(time_str):
    time_str = str(time_str)
    time_obj = datetime.strptime(time_str, '%H:%M:%S').time()
    return time_obj.strftime('%I:%M:%S %p')


# Function to calculate time difference
def calculate_time_difference(time_out, time_in):
    time_out_obj = datetime.strptime(time_out, '%I:%M:%S %p')
    time_in_obj = datetime.strptime(time_in, '%I:%M:%S %p')
    diff = time_out_obj - time_in_obj
    return diff


# Function to add total time diff per eid to the data frame
def append_summary(track_eid, df, last_row):
    summary_rows = []
    summary_rows.append({'Status': '', 'E. ID': '', 'Date': '', 'Time': '', 'Difference(O-I)': ''}) ## empty row
    summary_rows.append({'Status': '', 'E. ID': '', 'Date': 'Total', 'Time': '', 'Difference(O-I)': ''}) ## header "Total"
    summary_rows.append({'Status': '', 'E. ID': 'EID', 'Date': 'Date', 'Time': '', 'Difference(O-I)': 'Difference'})

    for eid, total_diff in track_eid.items():
        summary_rows.append({'Status': '', 'E. ID': eid, 'Date': last_row["Date"], 'Time': '', 'Difference(O-I)': format_time_difference(total_diff)})
        
    summary = pd.DataFrame(summary_rows)

    df = pd.concat([df, summary], ignore_index=True)
    return df


# Function to format time difference in AM/PM format
def format_time_difference(diff):
    seconds = diff.total_seconds()
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    time_str = '{:02}:{:02}:{:02}'.format(int(hours), int(minutes), int(seconds))
    print(f"time_str: {time_str}")
    return time_str


def modify_exel(exel_file):
    wb = load_workbook(exel_file)
    ws = wb.active

    # Center align all cells in the worksheet
    center_alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    # # Apply bold font and a different color to the last row
    # bold_font_for_total = Font(bold=True, color="FF0000")
    # bold_font = Font(bold=True)

    # for cell in ws[0]:
    #     cell.font = bold_font_for_total

    # for cell in ws[0]:
    #     cell.font = bold_font

    wb.save(exel_file)



def process_dataframe(df):
    ### sort the dataframe based on eid, date, time

    df['Date'] = pd.to_datetime(df['Date'], format='%Y/%m/%d')
    df['Time'] = pd.to_datetime(df['Time'], format='%H:%M:%S').dt.time

    df_sorted = df.sort_values(by=['E. ID', 'Date'])

    # df_sorted.to_csv('output.txt', sep='\t', index=False)

    rows = []

    ### add missing data
    for i in range(0, len(df_sorted) - 1):
        rows.append(df_sorted.iloc[i])
        ## conditions
        status1 = df_sorted.iloc[i]["Status"]
        status2 = df_sorted.iloc[i + 1]["Status"]
        if status1 == status2:
            status = 'I' if status1 == 'O' else 'O'
            if status == 'I':
                datetime_obj = datetime.combine(df_sorted.iloc[i]["Date"], df_sorted.iloc[i]["Time"])
                new_time = datetime_obj - timedelta(seconds=1)  # minus 1 second
                date = df_sorted.iloc[i]["Date"].strftime('%Y-%m-%d')
                new_row = pd.Series({
                    'Status': status,
                    'E. ID': df_sorted.iloc[i]['E. ID'],
                    'Date': df_sorted.iloc[i]["Date"],
                    'Time': new_time.strftime('%H:%M:%S'),
                })
                rows.append(new_row)
            else:
                datetime_obj = datetime.combine(df_sorted.iloc[i + 1]["Date"], df_sorted.iloc[i + 1]["Time"])
                new_time = datetime_obj + timedelta(seconds=1)  # add 1 second
                date = df_sorted.iloc[i + 1]["Date"].strftime('%Y-%m-%d')
                new_row = pd.Series({
                    'Status': status,
                    'E. ID': df_sorted.iloc[i + 1]['E. ID'],
                    'Date': df_sorted.iloc[i + 1]["Date"],
                    'Time': new_time.strftime('%H:%M:%S'),
                })
                rows.append(new_row)
        else:
            if status1 == 'O' and status2 == 'I':
                date_out = df_sorted.iloc[i]["Date"].strftime('%Y-%m-%d')
                date_in = df_sorted.iloc[i + 1]["Date"].strftime('%Y-%m-%d')

                if date_out != date_in:
                    datetime_obj_out = datetime.combine(df_sorted.iloc[i]["Date"], df_sorted.iloc[i]["Time"])
                    datetime_obj_in = datetime.combine(df_sorted.iloc[i + 1]["Date"], df_sorted.iloc[i + 1]["Time"])
                    new_time_in = datetime_obj_out - timedelta(seconds=1)  # minus 1 second
                    new_time_out = datetime_obj_in + timedelta(seconds=1)  # add 1 second
                    
                    ## append login
                    new_row = pd.Series({
                        'Status': 'I',
                        'E. ID': df_sorted.iloc[i]['E. ID'],
                        'Date': df_sorted.iloc[i]["Date"],
                        'Time': new_time_in.strftime('%H:%M:%S'),
                    })
                    rows.append(new_row)

                    ## append logout
                    new_row = pd.Series({
                        'Status': 'O',
                        'E. ID': df_sorted.iloc[i + 1]['E. ID'],
                        'Date': df_sorted.iloc[i + 1]["Date"],
                        'Time': new_time_out.strftime('%H:%M:%S'),
                    })
                    rows.append(new_row)

    # Create a new DataFrame with the modified rows
    df_modified = pd.DataFrame(rows)

    # df_modified.to_csv('output2.txt', sep='\t', index=False)

    # print(df_sorted)
    return df_modified

