
import sys
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QLabel, QLineEdit,
                             QPushButton, QTableWidget, QTableWidgetItem, QFileDialog)
from PyQt5.QtCore import Qt


data_length = []


# Function to convert 24-hour time to 12-hour AM/PM format
def convert_time_format(time_str):
    time_obj = datetime.strptime(time_str, '%H:%M:%S')
    return time_obj.strftime('%I:%M:%S %p')


# Function to calculate time difference
def calculate_time_difference(time_out, time_in):
    time_out_obj = datetime.strptime(time_out, '%I:%M:%S %p')
    time_in_obj = datetime.strptime(time_in, '%I:%M:%S %p')
    diff = time_out_obj - time_in_obj
    return diff


# Function to add total time diff per eid to the data frame
def append_summary(track_eid, df, last_row):
    summary_rows = []
    summary_rows.append({'Status': '', 'E. ID': '', 'Date': '', 'Time': '', 'Difference(O-I)': ''}) ## empty row
    summary_rows.append({'Status': '', 'E. ID': '', 'Date': 'Total', 'Time': '', 'Difference(O-I)': ''}) ## header "Total"
    summary_rows.append({'Status': '', 'E. ID': 'EID', 'Date': 'Date', 'Time': '', 'Difference(O-I)': 'Difference'})

    for eid, total_diff in track_eid.items():
        summary_rows.append({'Status': '', 'E. ID': eid, 'Date': last_row["Date"], 'Time': '', 'Difference(O-I)': format_time_difference(total_diff)})
        
    summary = pd.DataFrame(summary_rows)

    df = pd.concat([df, summary], ignore_index=True)
    return df

# Function to format time difference in AM/PM format
def format_time_difference(diff):
    seconds = diff.total_seconds()
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    time_str = '{:02}:{:02}:{:02}'.format(int(hours), int(minutes), int(seconds))
    print(f"time_str: {time_str}")
    return time_str


def modify_exel(exel_file):
    wb = load_workbook(exel_file)
    ws = wb.active

    # Center align all cells in the worksheet
    center_alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    # Apply bold font and a different color to the last row
    bold_font_for_total = Font(bold=True, color="FF0000")
    bold_font = Font(bold=True)

    for cell in ws[data_length[0] + 3]:
        cell.font = bold_font_for_total

    for cell in ws[data_length[0] + 4]:
        cell.font = bold_font

    wb.save(exel_file)
    


def process_file(input_file):
    data = []
    with open(input_file, 'r') as file:
        for line in file:
            parts = line.strip().split(',')
            data.append(parts)

    # Create a DataFrame from the data
    df = pd.DataFrame(data, columns=['Status', 'E. ID', 'Date', 'Time'])

    # Calculate time differences and format time
    df['Time'] = df['Time'].apply(lambda x: convert_time_format(x))
    df['Difference(O-I)'] = ''

    data_length.append(len(df))

    ### track eid
    track_eid = defaultdict(lambda: timedelta())



    for i in range(0, len(df) - 1, 2):
        eid = df.at[i, "E. ID"]

        time_out = df.at[i, 'Time']
        time_in = df.at[i + 1, 'Time']
        diff = calculate_time_difference(time_out, time_in)
        df.at[i, 'Difference(O-I)'] = str(diff)

        #### update eid and total diff value
        track_eid[eid] += diff
        # total_diff += diff

    #### last row data
    last_row = df.iloc[-1]

    df = append_summary(track_eid, df, last_row)

    return df



class App(QWidget):
    def __init__(self):
        super().__init__()
        self.title = 'File Processor'
        self.initUI()

    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(100, 100, 800, 600)

        self.layout = QVBoxLayout()

        self.label = QLabel('Enter file name:', self)
        self.layout.addWidget(self.label)

        self.file_name_input = QLineEdit(self)
        self.layout.addWidget(self.file_name_input)

        self.browse_button = QPushButton('Browse', self)
        self.browse_button.clicked.connect(self.open_file_dialog)
        self.layout.addWidget(self.browse_button)

        self.process_button = QPushButton('Process File', self)
        self.process_button.clicked.connect(self.process_file)
        self.layout.addWidget(self.process_button)

        self.save_button = QPushButton('Save to Excel', self)
        self.save_button.clicked.connect(self.save_to_excel)
        self.save_button.setEnabled(False)  # Disabled until a file is processed
        self.layout.addWidget(self.save_button)

        self.table = QTableWidget(self)
        self.layout.addWidget(self.table)

        self.setLayout(self.layout)
        self.df = None  # To store the processed DataFrame

    def open_file_dialog(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Text File", "", "Text Files (*.txt);;All Files (*)", options=options)
        if file_name:
            self.file_name_input.setText(file_name)

    def process_file(self):
        file_name = self.file_name_input.text()
        if file_name:
            try:
                self.df = process_file(file_name)
                self.display_output(self.df)
                self.save_button.setEnabled(True)  # Enable the save button after processing
            except Exception as e:
                self.display_error(f"Error processing file: {str(e)}")

    def display_output(self, df):
        self.table.setColumnCount(len(df.columns))
        self.table.setRowCount(len(df))
        self.table.setHorizontalHeaderLabels(df.columns)

        for i in range(len(df)):
            for j in range(len(df.columns)):
                item = QTableWidgetItem(str(df.iat[i, j]))
                item.setTextAlignment(Qt.AlignCenter)  # Center-align the text
                self.table.setItem(i, j, item)
                
    def save_to_excel(self):
        if self.df is not None:
            options = QFileDialog.Options()
            save_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
            if save_path:
                if not save_path.endswith('.xlsx'):
                    save_path += '.xlsx'
                self.df.to_excel(save_path, index=False)
                modify_exel(save_path)
                self.display_message(f"File saved successfully at {save_path}")

    def display_error(self, message):
        self.table.clear()
        self.table.setRowCount(1)
        self.table.setColumnCount(1)
        self.table.setItem(0, 0, QTableWidgetItem(message))

    def display_message(self, message):
        self.table.clear()
        self.table.setRowCount(1)
        self.table.setColumnCount(1)
        self.table.setItem(0, 0, QTableWidgetItem(message))


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    ex.show()
    sys.exit(app.exec_())